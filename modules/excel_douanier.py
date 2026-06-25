import io
import os
import re
from collections import defaultdict, deque
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from utils.database import (
    calcule_et_sauvegarde_recap,
    delete_excel_lignes,
    get_excel_lignes,
    get_excel_recap,
    init_database,
    save_excel_lignes,
    save_packing_data,
)


DOSSIER_EXCEL_ID = 0


def arrondir_valeur(val, decimales=2):
    """Arrondi propre avec Decimal pour eviter les derives IEEE 754."""
    try:
        d = Decimal(str(val if val is not None else 0))
        quantize_str = Decimal("0." + "0" * decimales) if decimales > 0 else Decimal("1")
        return float(d.quantize(quantize_str, rounding=ROUND_HALF_UP))
    except (InvalidOperation, TypeError, ValueError):
        return 0.0


def _to_decimal(value) -> Decimal:
    if pd.isna(value):
        return Decimal("0")
    if isinstance(value, str):
        cleaned = (
            value.replace("\xa0", " ")
            .replace("$", "")
            .replace("EUR", "")
            .replace("USD", "")
            .replace("MAD", "")
            .strip()
        )
        cleaned = re.sub(r"\s+", "", cleaned)
        if "," in cleaned and "." not in cleaned:
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    else:
        cleaned = str(value)
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _norm_col(value) -> str:
    value = re.sub(r"__\d+$", "", str(value))
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _norm_match(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _cell_text(value) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_nombre(val):
    """
    Convertit n'importe quelle valeur de cellule Excel en float propre.
    Gere : float, int, str avec virgule/espace/unites, NaN, None.
    """
    if val is None:
        return 0.0
    import math

    try:
        if isinstance(val, (int, float)):
            if math.isnan(val) or math.isinf(val):
                return 0.0
            return float(val)

        s = str(val).strip()
        s = s.replace("\xa0", "").replace(" ", "")
        s = s.replace(",", ".")
        s = "".join(c for c in s if c.isdigit() or c == ".")
        return float(s) if s else 0.0
    except (ValueError, TypeError):
        return 0.0


def est_ligne_total(desi_val):
    if not desi_val:
        return False
    mots_total = [
        "total",
        "sous-total",
        "subtotal",
        "sum",
        "grand total",
        "totaux",
        "total general",
        "合计",
        "总计",
    ]
    return any(mot in str(desi_val).lower() for mot in mots_total)


def _row_contains_total(row):
    return any(est_ligne_total(_cell_text(value)) for value in row.tolist())


def _hs_text(value) -> str:
    text = _cell_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text.strip()


def detecter_colonne(df, variantes):
    colonnes_norm = {}
    for col in df.columns:
        colonnes_norm.setdefault(_norm_col(col), []).append(col)
    for variante in variantes:
        key = _norm_col(variante)
        if key in colonnes_norm:
            return max(
                colonnes_norm[key],
                key=lambda col: df[col].apply(lambda value: _cell_text(value) != "").sum(),
            )
    return None


def _make_unique_headers(values):
    headers = []
    counts = {}
    previous = ""
    for idx, value in enumerate(values):
        header = _cell_text(value)
        if not header and _norm_col(previous) in {"totalamount"}:
            header = previous
        if not header:
            header = f"Column {idx + 1}"

        counts[header] = counts.get(header, 0) + 1
        headers.append(header if counts[header] == 1 else f"{header}__{counts[header]}")
        if _cell_text(value):
            previous = header
    return headers


def _row_header_score(row, required_groups):
    headers = {_norm_col(value) for value in row}
    return sum(
        1
        for variantes in required_groups
        if any(_norm_col(variante) in headers for variante in variantes)
    )


def _prepare_table(df, required_groups):
    """Detecte la ligne d'entete reelle dans les feuilles commerciales."""
    if df.empty:
        return df

    best_idx = None
    best_score = -1
    max_rows = min(len(df), 80)
    for idx in range(max_rows):
        score = _row_header_score(df.iloc[idx].tolist(), required_groups)
        if score > best_score:
            best_idx = idx
            best_score = score

    if best_idx is None or best_score <= 0:
        return df

    prepared = df.iloc[best_idx + 1:].copy()
    prepared.columns = _make_unique_headers(df.iloc[best_idx].tolist())
    prepared = prepared.dropna(how="all").reset_index(drop=True)
    return prepared


def lire_feuille_invoice(df):
    df = _prepare_table(
        df,
        [
            ["Qty", "Quantity", "QTE", "Quantite", "QTY"],
            ["Description", "Designation", "DESI", "Item", "Description of Goods"],
        ],
    )
    qte_col = detecter_colonne(df, ["Qty", "Quantity", "QTE", "Quantite", "QTY"])
    desi_col = detecter_colonne(
        df, ["Description", "Designation", "DESI", "Item", "Description of Goods"]
    )
    hs_col = detecter_colonne(df, ["HS#", "HS Code", "HS", "Code SH", "Code_SH", "HSCode"])
    unit_col = detecter_colonne(df, ["Unit Price", "Prix Unitaire", "Unit_Price"])
    total_col = detecter_colonne(df, ["Total Amount", "Total", "Valeur", "Amount", "Total_Amount"])
    pn_col = detecter_colonne(df, ["Materials", "PN", "Part Number", "Part No", "Ref", "Material"])

    if not desi_col or not qte_col:
        raise ValueError("colonnes Invoice minimales introuvables (QTE/DESI)")

    lignes = []
    for _, row in df.iterrows():
        desi = _cell_text(row.get(desi_col))
        qte_dec = _to_decimal(row.get(qte_col))
        if not desi or qte_dec == 0:
            continue

        if total_col:
            val_dec = _to_decimal(row.get(total_col))
        else:
            val_dec = qte_dec * _to_decimal(row.get(unit_col)) if unit_col else Decimal("0")

        lignes.append({
            "qte": arrondir_valeur(qte_dec, 3),
            "desi": desi,
            "hs": _hs_text(row.get(hs_col)) if hs_col else "",
            "val": arrondir_valeur(val_dec, 2),
            "pn": _cell_text(row.get(pn_col)) if pn_col else "",
        })
    return lignes


def lire_feuille_packing(df):
    df = _prepare_table(
        df,
        [
            ["Description", "Designation", "DESI", "Description of Goods"],
            ["Net Weight", "Poids Net", "Net_Weight", "NetWeight"],
            ["Gross Weight", "Poids Brut", "Gross_Weight"],
        ],
    )
    qte_col = detecter_colonne(df, ["Qty", "Quantity", "Quantities", "QTE"])
    desi_col = detecter_colonne(df, ["Description", "Designation", "DESI", "Description of Goods"])
    net_col = detecter_colonne(df, ["Net Weight", "Poids Net", "Net_Weight", "NetWeight"])
    gross_col = detecter_colonne(df, ["Gross Weight", "Poids Brut", "Gross_Weight"])
    hs_col = detecter_colonne(df, ["HS#", "HS Code", "HS", "Code SH", "Code_SH", "HSCode"])
    pn_col = detecter_colonne(df, ["Materials", "PN", "Part Number", "Material"])

    if not desi_col:
        raise ValueError("colonne Packing designation introuvable")

    total_markers = [idx for idx, row in df.iterrows() if _row_contains_total(row)]
    if total_markers:
        cutoff = total_markers[0]
        df = df.iloc[:cutoff].copy()
        if not df.empty:
            last_idx = df.index[-1]
            last_row = df.loc[last_idx]
            last_desi = _cell_text(last_row.get(desi_col))
            last_pn = _cell_text(last_row.get(pn_col)) if pn_col else ""
            last_qte = parse_nombre(last_row.get(qte_col)) if qte_col else 0.0
            last_poids_net = parse_nombre(last_row.get(net_col)) if net_col else 0.0
            if not last_desi and not last_pn and (last_qte != 0 or last_poids_net != 0):
                df = df.drop(index=last_idx)

    colonnes_a_ffill = [desi_col, hs_col, pn_col]
    for col in colonnes_a_ffill:
        if col and col in df.columns:
            with pd.option_context("future.no_silent_downcasting", True):
                df[col] = df[col].ffill()

    lignes = []
    for _, row in df.iterrows():
        desi = _cell_text(row.get(desi_col))
        qte = parse_nombre(row.get(qte_col)) if qte_col else 0.0
        poids_net = parse_nombre(row.get(net_col)) if net_col else 0.0
        poids_brut = parse_nombre(row.get(gross_col)) if gross_col else 0.0

        if est_ligne_total(desi):
            continue

        if not desi and poids_net == 0 and qte == 0:
            continue

        lignes.append({
            "qte": arrondir_valeur(qte, 3),
            "desi": desi,
            "poids_net": arrondir_valeur(poids_net, 3),
            "poids_brut": arrondir_valeur(poids_brut, 3),
            "hs": _hs_text(row.get(hs_col)) if hs_col else "",
            "pn": _cell_text(row.get(pn_col)) if pn_col else "",
        })
    total_pn = sum(ligne["poids_net"] for ligne in lignes)
    st.write(f"  └─ Poids net total lu : {total_pn:.3f} kg ({len(lignes)} lignes)")
    return lignes


def parser_fichier_excel(fichier_bytes, nom_fichier):
    suffix = os.path.splitext(nom_fichier)[1].lower()
    engine = "xlrd" if suffix == ".xls" else "openpyxl"
    result = {
        "source": nom_fichier,
        "lignes_invoice": [],
        "lignes_packing": [],
        "erreurs": [],
    }

    try:
        feuilles = pd.read_excel(
            io.BytesIO(fichier_bytes),
            sheet_name=None,
            engine=engine,
            header=None,
        )
    except Exception as e:
        result["erreurs"].append(f"lecture Excel impossible : {e}")
        return result

    invoice_name = next(
        (name for name in feuilles if "invoice" in name.lower() or "facture" in name.lower()),
        None,
    )
    packing_name = next((name for name in feuilles if "packing" in name.lower()), None)

    if invoice_name:
        try:
            result["lignes_invoice"] = lire_feuille_invoice(feuilles[invoice_name])
        except Exception as e:
            result["erreurs"].append(f"Invoice : {e}")
    else:
        result["erreurs"].append("feuille Invoice introuvable")

    if packing_name:
        try:
            result["lignes_packing"] = lire_feuille_packing(feuilles[packing_name])
        except Exception as e:
            result["erreurs"].append(f"Packing List : {e}")
    else:
        result["erreurs"].append("feuille Packing List introuvable")

    return result


def enrichir_avec_packing(lignes_invoice, lignes_packing):
    by_pn = defaultdict(deque)
    by_desc = defaultdict(deque)
    for idx, ligne in enumerate(lignes_packing):
        pn_key = _norm_match(ligne.get("pn"))
        desc_key = _norm_match(ligne.get("desi"))
        if pn_key:
            by_pn[pn_key].append(idx)
        if desc_key:
            by_desc[desc_key].append(idx)

    used_packing = set()

    def take_match(queue):
        while queue:
            idx = queue.popleft()
            if idx not in used_packing:
                used_packing.add(idx)
                return lignes_packing[idx]
        return None

    lignes = []
    for ligne in lignes_invoice:
        match = None
        pn_key = _norm_match(ligne.get("pn"))
        desc_key = _norm_match(ligne.get("desi"))
        if pn_key:
            match = take_match(by_pn[pn_key])
        if match is None and desc_key:
            match = take_match(by_desc[desc_key])

        enriched = dict(ligne)
        enriched["poids_net"] = arrondir_valeur(match.get("poids_net", 0), 3) if match else 0
        enriched["poids_brut"] = arrondir_valeur(match.get("poids_brut", 0), 3) if match else 0
        if not enriched.get("hs") and match and match.get("hs"):
            enriched["hs"] = match["hs"]
        lignes.append(enriched)
    return lignes


def show_excel_douanier():
    if "user" not in st.session_state:
        st.error("Vous devez etre connecte.")
        return

    try:
        init_database()
    except Exception as e:
        st.error(f"Erreur d'initialisation SQLite : {e}")
        return

    st.markdown(
        """
        <div class="page-header">
            <h2> Recapitulatif Douanier Excel</h2>
            <p>Importez plusieurs fichiers Excel, fusionnez Feuil2 et exportez Feuil4</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.subheader("Importer des fichiers Excel")
    fichiers = st.file_uploader(
        "Glisser-deposer vos fichiers Excel ici",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help=(
            "Vous pouvez uploader plusieurs fichiers en meme temps. "
            "Chaque fichier doit contenir une feuille Invoice et/ou Packing List."
        ),
    )

    mode_import = st.radio(
        "Mode d'import",
        ["Ajouter aux lignes existantes", "Remplacer les lignes existantes"],
        horizontal=True,
    )

    if fichiers and st.button("Importer et fusionner", type="primary"):
        toutes_lignes = []
        toutes_packing = []
        statuses = []

        for fichier in fichiers:
            parsed = parser_fichier_excel(fichier.getvalue(), fichier.name)
            packing = [
                {**ligne, "source_fichier": fichier.name}
                for ligne in parsed["lignes_packing"]
            ]
            invoice = [
                {**ligne, "source_fichier": fichier.name}
                for ligne in parsed["lignes_invoice"]
            ]
            lignes_enrichies = enrichir_avec_packing(invoice, packing)

            toutes_lignes.extend(lignes_enrichies)
            toutes_packing.extend(packing)

            if parsed["erreurs"] and not parsed["lignes_invoice"]:
                statuses.append(
                    f"❌ {fichier.name} -> Erreur: {'; '.join(parsed['erreurs'])}"
                )
            else:
                statuses.append(
                    f"✅ {fichier.name} -> Invoice: {len(parsed['lignes_invoice'])} lignes | "
                    f"Packing: {len(parsed['lignes_packing'])} lignes"
                )

        for status in statuses:
            st.write(status)

        try:
            replace = mode_import.startswith("Remplacer")
            save_excel_lignes(DOSSIER_EXCEL_ID, toutes_lignes, replace=replace)
            save_packing_data(DOSSIER_EXCEL_ID, toutes_packing)
            calcule_et_sauvegarde_recap(DOSSIER_EXCEL_ID)
            st.success(f"{len(toutes_lignes)} lignes importees depuis {len(fichiers)} fichiers")
            st.rerun()
        except Exception as e:
            st.error(f"Erreur lors de l'import : {e}")

    st.markdown("---")
    st.markdown("###  Feuil2 - Detail des marchandises")

    try:
        lignes = get_excel_lignes(DOSSIER_EXCEL_ID)
    except Exception as e:
        st.error(f"Erreur lors du chargement des lignes : {e}")
        lignes = []

    df = pd.DataFrame(
        lignes,
        columns=["id", "source_fichier", "qte", "desi", "hs", "val", "poids_net", "pn"],
    )
    if df.empty:
        df_display = pd.DataFrame(columns=["SOURCE", "QTE", "DESI", "HS", "VAL", "POIDS_NET"])
    else:
        df_display = df[["source_fichier", "qte", "desi", "hs", "val", "poids_net"]].copy()
        df_display.columns = ["SOURCE", "QTE", "DESI", "HS", "VAL", "POIDS_NET"]
        df_display["QTE"] = df_display["QTE"].apply(lambda x: arrondir_valeur(x, 3))
        df_display["VAL"] = df_display["VAL"].apply(lambda x: arrondir_valeur(x, 2))
        df_display["POIDS_NET"] = df_display["POIDS_NET"].apply(lambda x: arrondir_valeur(x, 3))

    edited_df = st.data_editor(
        df_display,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "SOURCE": st.column_config.TextColumn("Source", width="small"),
            "QTE": st.column_config.NumberColumn("QTE", min_value=0, format="%.3f"),
            "DESI": st.column_config.TextColumn("Designation", width="large"),
            "HS": st.column_config.TextColumn("Code HS", width="medium"),
            "VAL": st.column_config.NumberColumn("Valeur", min_value=0, format="%.2f"),
            "POIDS_NET": st.column_config.NumberColumn("Poids net", min_value=0, format="%.3f"),
        },
        key="feuil2_editor",
    )

    col1, col2 = st.columns([1, 1])
    with col1:
        if st.button(" Sauvegarder Feuil2", type="primary"):
            try:
                rows = []
                for _, row in edited_df.iterrows():
                    desi = _cell_text(row.get("DESI"))
                    hs = _hs_text(row.get("HS"))
                    if not desi or not hs:
                        continue
                    rows.append({
                        "source_fichier": _cell_text(row.get("SOURCE")),
                        "qte": arrondir_valeur(row.get("QTE"), 3),
                        "desi": desi,
                        "hs": hs,
                        "val": arrondir_valeur(row.get("VAL"), 2),
                        "poids_net": arrondir_valeur(row.get("POIDS_NET"), 3),
                        "pn": "",
                    })
                save_excel_lignes(DOSSIER_EXCEL_ID, rows, replace=True)
                calcule_et_sauvegarde_recap(DOSSIER_EXCEL_ID)
                st.success("Feuil2 sauvegardee")
                st.rerun()
            except Exception as e:
                st.error(f"Erreur lors de la sauvegarde : {e}")

    with col2:
        if st.button("🗑 Vider tout", type="secondary"):
            st.session_state["confirm_vider_excel"] = True

    if st.session_state.get("confirm_vider_excel"):
        st.warning("Confirmer la suppression de toutes les lignes ?")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Oui, vider", key="confirm_oui_excel"):
                delete_excel_lignes(DOSSIER_EXCEL_ID)
                st.session_state["confirm_vider_excel"] = False
                st.rerun()
        with c2:
            if st.button("Annuler", key="confirm_non_excel"):
                st.session_state["confirm_vider_excel"] = False
                st.rerun()

    st.markdown("---")
    st.markdown("###  Feuil4 - Recapitulatif par code HS")

    try:
        recap = calcule_et_sauvegarde_recap(DOSSIER_EXCEL_ID)
    except Exception as e:
        st.error(f"Erreur lors du calcul du recapitulatif : {e}")
        recap = get_excel_recap(DOSSIER_EXCEL_ID)

    if recap:
        df_recap = pd.DataFrame(recap)[
            ["hs", "desi", "qte_total", "val_total", "poids_net_total"]
        ]
        df_recap.columns = ["HS", "DESI", "QTE_TOTAL", "VAL_TOTAL", "POIDS_NET_TOTAL"]
        df_recap["QTE_TOTAL"] = df_recap["QTE_TOTAL"].apply(lambda x: arrondir_valeur(x, 3))
        df_recap["VAL_TOTAL"] = df_recap["VAL_TOTAL"].apply(lambda x: arrondir_valeur(x, 2))
        df_recap["POIDS_NET_TOTAL"] = df_recap["POIDS_NET_TOTAL"].apply(
            lambda x: arrondir_valeur(x, 3)
        )
        st.dataframe(df_recap, use_container_width=True, hide_index=True)

        total_feuil2 = sum(Decimal(str(l.get("val") or 0)) for l in lignes)
        total_feuil4 = sum(Decimal(str(r.get("val_total") or 0)) for r in recap)
        ecart = abs(total_feuil2 - total_feuil4)
        c1, c2, c3 = st.columns(3)
        c1.metric("Total VAL Feuil2", f"{arrondir_valeur(total_feuil2, 2):,.2f}")
        c2.metric("Total VAL Feuil4", f"{arrondir_valeur(total_feuil4, 2):,.2f}")
        if ecart < Decimal("0.01"):
            c3.success("Totaux coherents")
        else:
            c3.error(f"Ecart de {arrondir_valeur(ecart, 2):,.2f}")
    else:
        st.info("Aucune donnee. Importez des fichiers Excel ou ajoutez des lignes dans Feuil2.")

    st.markdown("---")
    if lignes:
        st.download_button(
            label="📥 Telecharger Excel (.xlsx)",
            data=generate_excel(lignes, recap),
            file_name="recap_douanier.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
    else:
        st.info("Ajoutez des lignes dans Feuil2 pour activer l'export.")


def _style_header(cell, fill):
    cell.fill = fill
    cell.font = Font(bold=True, color="FFFFFF", name="Poppins")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _center_cell(cell, wrap_text=False):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap_text)


def generate_excel(lignes: list[dict], recap: list[dict]) -> bytes:
    wb = Workbook()
    bleu_header = PatternFill("solid", fgColor="4472C4")
    vert_header = PatternFill("solid", fgColor="1F7A4A")
    jaune_data = PatternFill("solid", fgColor="FFFF00")
    orange_total = PatternFill("solid", fgColor="FFA500")
    gris_clair = PatternFill("solid", fgColor="F2F2F2")
    blanc = PatternFill("solid", fgColor="FFFFFF")
    font_total = Font(bold=True, name="Poppins")

    ws2 = wb.active
    ws2.title = "Feuil2"
    headers2 = ["QTE", "DESI", "HS", "VAL", "POIDS NET"]
    for col_idx, titre in enumerate(headers2, 1):
        _style_header(ws2.cell(row=1, column=col_idx, value=titre), bleu_header)

    for row_idx, ligne in enumerate(lignes, 2):
        values = [
            arrondir_valeur(ligne.get("qte"), 3),
            ligne.get("desi", ""),
            str(ligne.get("hs", "") or ""),
            arrondir_valeur(ligne.get("val"), 2),
            arrondir_valeur(ligne.get("poids_net"), 3),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = jaune_data
            _center_cell(cell, wrap_text=(col_idx == 2))
            if col_idx == 1 or col_idx == 5:
                cell.number_format = "#,##0.###"
            if col_idx == 4:
                cell.number_format = "#,##0.00"

    total_row2 = len(lignes) + 2
    ws2.cell(row=total_row2, column=1, value="TOTAL").fill = orange_total
    ws2.cell(row=total_row2, column=1).font = font_total
    _center_cell(ws2.cell(row=total_row2, column=1))
    total_val2 = sum(Decimal(str(l.get("val") or 0)) for l in lignes)
    total_poids2 = sum(Decimal(str(l.get("poids_net") or 0)) for l in lignes)
    ws2.cell(row=total_row2, column=4, value=arrondir_valeur(total_val2, 2)).fill = orange_total
    ws2.cell(row=total_row2, column=4).font = font_total
    _center_cell(ws2.cell(row=total_row2, column=4))
    ws2.cell(row=total_row2, column=4).number_format = "#,##0.00"
    ws2.cell(row=total_row2, column=5, value=arrondir_valeur(total_poids2, 3)).fill = orange_total
    ws2.cell(row=total_row2, column=5).font = font_total
    _center_cell(ws2.cell(row=total_row2, column=5))
    ws2.cell(row=total_row2, column=5).number_format = "#,##0.###"

    widths2 = [10, 45, 16, 14, 14]
    for idx, width in enumerate(widths2, 1):
        ws2.column_dimensions[chr(64 + idx)].width = width
    ws2.row_dimensions[1].height = 22

    ws4 = wb.create_sheet("Feuil4")
    headers4 = ["QTE", "DESI", "HS", "VAL", "POIDS NET"]
    for col_idx, titre in enumerate(headers4, 1):
        _style_header(ws4.cell(row=1, column=col_idx, value=titre), vert_header)

    for row_idx, rec in enumerate(recap, 2):
        fond = gris_clair if row_idx % 2 == 0 else blanc
        values = [
            arrondir_valeur(rec.get("qte_total"), 3),
            rec.get("desi", ""),
            str(rec.get("hs", "") or ""),
            arrondir_valeur(rec.get("val_total"), 2),
            arrondir_valeur(rec.get("poids_net_total"), 3),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fond
            _center_cell(cell, wrap_text=(col_idx == 2))
            if col_idx in (1, 5):
                cell.number_format = "#,##0.###"
            if col_idx == 4:
                cell.number_format = "#,##0.00"

    total_row4 = len(recap) + 2
    ws4.cell(row=total_row4, column=1, value="TOTAL").fill = orange_total
    ws4.cell(row=total_row4, column=1).font = font_total
    _center_cell(ws4.cell(row=total_row4, column=1))
    total_val4 = sum(Decimal(str(r.get("val_total") or 0)) for r in recap)
    total_poids4 = sum(Decimal(str(r.get("poids_net_total") or 0)) for r in recap)
    ws4.cell(row=total_row4, column=4, value=arrondir_valeur(total_val4, 2)).fill = orange_total
    ws4.cell(row=total_row4, column=4).font = font_total
    _center_cell(ws4.cell(row=total_row4, column=4))
    ws4.cell(row=total_row4, column=4).number_format = "#,##0.00"
    ws4.cell(row=total_row4, column=5, value=arrondir_valeur(total_poids4, 3)).fill = orange_total
    ws4.cell(row=total_row4, column=5).font = font_total
    _center_cell(ws4.cell(row=total_row4, column=5))
    ws4.cell(row=total_row4, column=5).number_format = "#,##0.###"

    widths4 = [10, 45, 16, 14, 14]
    for idx, width in enumerate(widths4, 1):
        ws4.column_dimensions[chr(64 + idx)].width = width
    ws4.row_dimensions[1].height = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
