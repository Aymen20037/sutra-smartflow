import streamlit as st
from utils.database import get_current_user_dossiers, get_dossier_by_ref, get_documents_by_dossier, get_articles_by_document
from utils.helpers import format_currency, calculate_total_weight, calculate_total_value, get_exchange_rates
import json
import io
import xml.etree.ElementTree as ET
from xml.dom import minidom
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def show_export():
    """Affiche l'interface d'export des données"""
    st.title(" Export des données pour les douanes")

    if 'current_dossier_id' not in st.session_state:
        st.warning("Aucun dossier actif. Veuillez d'abord téléverser et réviser des documents.")
        return

    dossier_id = st.session_state['current_dossier_id']
    ref_interne = st.session_state.get('current_ref_interne', 'N/A')
    allowed_dossiers = get_current_user_dossiers()
    allowed_ids = [d["id"] for d in allowed_dossiers]

    if dossier_id not in allowed_ids:
        st.error("Dossier non autorise pour cet utilisateur.")
        return

    st.info(f"Export du dossier : {ref_interne}")

    dossier = get_dossier_by_ref(ref_interne)
    if not dossier:
        st.error("Dossier introuvable.")
        return

    documents = get_documents_by_dossier(dossier_id)
    all_articles = []
    for doc in documents:
        articles = get_articles_by_document(doc['id'])
        for article in articles:
            article['document_id']     = doc['id']
            article['type_document']   = doc['type_document']
            article['numero_document'] = doc.get('numero_document', '')
            article['date_document']   = doc.get('date_document', '')
            article['devise']          = doc.get('devise', '')
            article['expediteur']      = {
                'nom':         doc.get('expediteur_nom', ''),
                'adresse':     doc.get('expediteur_adresse', ''),
                'ville':       doc.get('expediteur_ville', ''),
                'code_postal': doc.get('expediteur_code_postal', ''),
                'pays':        doc.get('expediteur_pays', '')
            }
            article['destinataire']    = {
                'nom':         doc.get('destinataire_nom', ''),
                'adresse':     doc.get('destinataire_adresse', ''),
                'ville':       doc.get('destinataire_ville', ''),
                'code_postal': doc.get('destinataire_code_postal', ''),
                'pays':        doc.get('destinataire_pays', '')
            }
        all_articles.extend(articles)

    if not all_articles:
        st.warning("Aucun article à exporter. Veuillez d'abord extraire et réviser des données.")
        return

    total_weight = calculate_total_weight(all_articles)
    financial_totals = get_document_financial_totals(documents, all_articles)
    total_value_doc = financial_totals["valeur_totale_doc"]
    total_value_mad = convert_currency(
        total_value_doc,
        financial_totals["devise"] or "USD",
        'MAD'
    )

    # ── Règles métiers ──────────────────────────────────────────────────────
    st.subheader("Vérification des règles métiers")
    col1, col2 = st.columns(2)

    with col1:
        if total_weight <= 0:
            st.error(" Le poids total est nul ou négatif. Veuillez vérifier les données.")
        elif total_weight > 2000:
            st.warning(f" Le poids total ({total_weight:.2f} kg) dépasse 2000 kg. Veuillez vérifier.")
        else:
            st.success(f" Poids total cohérent : {total_weight:.2f} kg")

    with col2:
        if total_value_doc <= 0:
            st.error(" La valeur totale est nulle ou négative. Veuillez vérifier les données.")
        else:
            st.success(
                " Total TTC cohérent : "
                f"{format_currency(total_value_doc, financial_totals['devise'])}"
            )

    # ── Récapitulatif ───────────────────────────────────────────────────────
    st.subheader("Récapitulatif de l'export")
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric(label="Documents",       value=len(documents))
    with col2:
        st.metric(label="Articles",         value=len(all_articles))
    with col3:
        st.metric(label="Poids total (kg)", value=f"{total_weight:.2f}")
    with col4:
        st.metric(
            label="Valeur totale MAD",
            value=format_currency(total_value_mad, "MAD")
        )

    fin_col1, fin_col2, fin_col3 = st.columns(3)
    with fin_col1:
        st.metric("Montant HT", format_currency(financial_totals["montant_ht"], financial_totals["devise"]))
    with fin_col2:
        st.metric("TVA", format_currency(financial_totals["montant_tva"], financial_totals["devise"]))
    with fin_col3:
        st.metric("Total TTC", format_currency(total_value_doc, financial_totals["devise"]))

    # ── Format & génération ─────────────────────────────────────────────────
    st.subheader("Génération du fichier d'export")
    export_format = st.radio(
        "Choisissez le format d'export :",
        options=["JSON", "XML", "Excel"],
        horizontal=True
    )

    if st.button("🚀 Générer le fichier d'export", type="primary"):
        try:
            if export_format == "JSON":
                export_data  = prepare_json_export(
                    dossier, documents, all_articles,
                    total_weight, total_value_doc, total_value_mad,
                    financial_totals
                )
                file_content = json.dumps(export_data, indent=2, ensure_ascii=False)
                file_name    = f"export_{ref_interne}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                mime_type    = "application/json"

            elif export_format == "XML":
                file_content = prepare_xml_export(
                    dossier, documents, all_articles,
                    total_weight, total_value_doc, total_value_mad,
                    financial_totals
                )
                file_name = f"export_{ref_interne}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml"
                mime_type = "application/xml"

            else:  # Excel
                file_content = prepare_xlsx_export(
                    dossier, documents, all_articles,
                    total_weight, total_value_doc, total_value_mad,
                    financial_totals
                )
                file_name = f"export_{ref_interne}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            downloaded = st.download_button(
                label=f" Télécharger l'export {export_format}",
                data=file_content,
                file_name=file_name,
                mime=mime_type
            )
            if downloaded:
                st.session_state["last_exported_dossier"] = dossier["id"]
            st.success(f"Fichier {export_format} généré avec succès !")

        except Exception as e:
            st.error(f"Erreur lors de la génération de l'export : {str(e)}")


# ═══════════════════════════════════════════════════════════════════════════════
# Fonctions d'export
# ═══════════════════════════════════════════════════════════════════════════════

def _to_float(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def get_document_financial_totals(documents, articles):
    devise = next((doc.get('devise', '') for doc in documents if doc.get('devise')), 'USD')
    montant_ht = sum(_to_float(doc.get('montant_ht', 0)) for doc in documents)
    montant_tva = sum(_to_float(doc.get('montant_tva', 0)) for doc in documents)
    valeur_totale_doc = sum(_to_float(doc.get('valeur_totale_doc', 0)) for doc in documents)

    if valeur_totale_doc <= 0:
        valeur_totale_doc = calculate_total_value(articles)

    return {
        "montant_ht": montant_ht,
        "montant_tva": montant_tva,
        "valeur_totale_doc": valeur_totale_doc,
        "devise": devise or "USD",
    }


def prepare_json_export(
    dossier,
    documents,
    articles,
    total_weight,
    total_value_doc,
    total_value_mad,
    financial_totals=None
):
    total_brut = sum(article.get('poids_brut', 0) for article in articles)

    first_doc = documents[0] if documents else {}
    financial_totals = financial_totals or get_document_financial_totals(documents, articles)

    export_data = {
        "numero":    first_doc.get('numero_document', ''),
        "date":      first_doc.get('date_document', ''),
        "expediteur": {
            "nom":         first_doc.get('expediteur_nom', ''),
            "adresse":     first_doc.get('expediteur_adresse', ''),
            "ville":       first_doc.get('expediteur_ville', ''),
            "code_postal": first_doc.get('expediteur_code_postal', ''),
            "pays":        first_doc.get('expediteur_pays', '')
        },
        "destinataire": {
            "nom":         first_doc.get('destinataire_nom', ''),
            "adresse":     first_doc.get('destinataire_adresse', ''),
            "ville":       first_doc.get('destinataire_ville', ''),
            "code_postal": first_doc.get('destinataire_code_postal', ''),
            "pays":        first_doc.get('destinataire_pays', '')
        },
        "montant_ht":        round(financial_totals["montant_ht"], 2),
        "montant_tva":       round(financial_totals["montant_tva"], 2),
        "valeur_totale_doc": round(financial_totals["valeur_totale_doc"], 2),
        "valeur_totale":     round(total_value_doc, 2),
        "devise":            financial_totals["devise"],
        "dossier": {
            "ref_interne":   dossier['ref_interne'],
            "statut":        dossier['statut'],
            "date_creation": dossier['date_creation']
        },
        "export_info": {
            "date_export": datetime.now().isoformat(),
            "version":     "1.0",
            "generateur":  "Système de Transit Douanier Streamlit"
        },
        "totaux": {
            "poids_net_kg":      round(total_weight, 2),
            "poids_brut_kg":     round(total_brut, 2),
            "montant_ht":        round(financial_totals["montant_ht"], 2),
            "montant_tva":       round(financial_totals["montant_tva"], 2),
            "total_ttc":         round(total_value_doc, 2),
            "devise":            financial_totals["devise"],
            "valeur_totale_mad": round(total_value_mad, 2)
        },
        "documents": [],
        "articles":  []
    }

    for doc in documents:
        export_data["documents"].append({
            "id":             doc['id'],
            "type":           doc['type_document'],
            "chemin_fichier": doc['chemin_fichier'],
            "date_import":    doc['date_import'],
            "montant_ht":     round(_to_float(doc.get('montant_ht', 0)), 2),
            "montant_tva":    round(_to_float(doc.get('montant_tva', 0)), 2),
            "total_ttc":      round(_to_float(doc.get('valeur_totale_doc', 0)), 2),
            "devise":         doc.get('devise', '')
        })

    for article in articles:
        q = article.get('quantite', 0)
        v = article.get('valeur_devise', 0)
        export_data["articles"].append({
            "num_ligne":       article.get('num_ligne'),
            "designation":     article.get('designation', ''),
            "code_sh":         article.get('code_sh', ''),
            "quantite":        q,
            "unite":           article.get('unite', ''),
            "poids_net":       round(article.get('poids_net', 0), 2),
            "poids_brut":      round(article.get('poids_brut', 0), 2),
            "valeur_unitaire": round(article.get('valeur_unitaire', 0), 2),
            "valeur_totale":   round(v, 2),
            "origine":         article.get('origine', '')
        })

    return export_data


def prepare_xml_export(
    dossier,
    documents,
    articles,
    total_weight,
    total_value_doc,
    total_value_mad,
    financial_totals=None
):
    total_brut = sum(article.get('poids_brut', 0) for article in articles)
    first_doc  = documents[0] if documents else {}
    financial_totals = financial_totals or get_document_financial_totals(documents, articles)

    root = ET.Element("EXPORT_DOUANIER_MAROC")

    # ── DOSSIER ─────────────────────────────────────────────────────────────
    d = ET.SubElement(root, "DOSSIER")
    ET.SubElement(d, "REF_INTERNE").text       = dossier['ref_interne']
    ET.SubElement(d, "STATUT").text            = dossier['statut']
    ET.SubElement(d, "DATE_CREATION").text     = str(dossier['date_creation'])
    ET.SubElement(d, "NUMERO_DOCUMENT").text   = first_doc.get('numero_document', '')
    ET.SubElement(d, "DATE_DOCUMENT").text     = first_doc.get('date_document', '')
    ET.SubElement(d, "DEVISE").text            = financial_totals["devise"]

    # ── EXPÉDITEUR ──────────────────────────────────────────────────────────
    exp_el = ET.SubElement(root, "EXPEDITEUR")
    ET.SubElement(exp_el, "NOM").text          = first_doc.get('expediteur_nom', '')
    ET.SubElement(exp_el, "ADRESSE").text      = first_doc.get('expediteur_adresse', '')
    ET.SubElement(exp_el, "VILLE").text        = first_doc.get('expediteur_ville', '')
    ET.SubElement(exp_el, "CODE_POSTAL").text  = first_doc.get('expediteur_code_postal', '')
    ET.SubElement(exp_el, "PAYS").text         = first_doc.get('expediteur_pays', '')

    # ── DESTINATAIRE ────────────────────────────────────────────────────────
    dest_el = ET.SubElement(root, "DESTINATAIRE")
    ET.SubElement(dest_el, "NOM").text         = first_doc.get('destinataire_nom', '')
    ET.SubElement(dest_el, "ADRESSE").text     = first_doc.get('destinataire_adresse', '')
    ET.SubElement(dest_el, "VILLE").text       = first_doc.get('destinataire_ville', '')
    ET.SubElement(dest_el, "CODE_POSTAL").text = first_doc.get('destinataire_code_postal', '')
    ET.SubElement(dest_el, "PAYS").text        = first_doc.get('destinataire_pays', '')

    # ── INFO EXPORT ─────────────────────────────────────────────────────────
    info = ET.SubElement(root, "INFO_EXPORT")
    ET.SubElement(info, "DATE_EXPORT").text    = datetime.now().isoformat()
    ET.SubElement(info, "VERSION").text        = "1.0"
    ET.SubElement(info, "GENERATEUR").text     = "Système de Transit Douanier Streamlit"

    # ── TOTAUX ──────────────────────────────────────────────────────────────
    tot = ET.SubElement(root, "TOTALS")
    ET.SubElement(tot, "POIDS_NET_KG").text      = f"{total_weight:.2f}"
    ET.SubElement(tot, "POIDS_BRUT_KG").text     = f"{total_brut:.2f}"
    ET.SubElement(tot, "MONTANT_HT").text        = f"{financial_totals['montant_ht']:.2f}"
    ET.SubElement(tot, "MONTANT_TVA").text       = f"{financial_totals['montant_tva']:.2f}"
    ET.SubElement(tot, "TOTAL_TTC").text         = f"{total_value_doc:.2f}"
    ET.SubElement(tot, "DEVISE").text            = financial_totals["devise"]
    ET.SubElement(tot, "VALEUR_TOTALE_MAD").text = f"{total_value_mad:.2f}"

    # ── DOCUMENTS ───────────────────────────────────────────────────────────
    docs_el = ET.SubElement(root, "DOCUMENTS")
    for doc in documents:
        de = ET.SubElement(docs_el, "DOCUMENT")
        ET.SubElement(de, "ID").text             = str(doc['id'])
        ET.SubElement(de, "TYPE").text           = doc['type_document']
        ET.SubElement(de, "CHEMIN_FICHIER").text = doc['chemin_fichier']
        ET.SubElement(de, "DATE_IMPORT").text    = str(doc['date_import'])
        ET.SubElement(de, "MONTANT_HT").text     = f"{_to_float(doc.get('montant_ht', 0)):.2f}"
        ET.SubElement(de, "MONTANT_TVA").text    = f"{_to_float(doc.get('montant_tva', 0)):.2f}"
        ET.SubElement(de, "TOTAL_TTC").text      = f"{_to_float(doc.get('valeur_totale_doc', 0)):.2f}"
        ET.SubElement(de, "DEVISE").text         = doc.get('devise', '')

    # ── ARTICLES ────────────────────────────────────────────────────────────
    arts_el = ET.SubElement(root, "ARTICLES")
    for article in articles:
        q  = article.get('quantite', 0)
        v  = article.get('valeur_devise', 0)
        ae = ET.SubElement(arts_el, "ARTICLE")
        ET.SubElement(ae, "DOCUMENT_ID").text        = str(article['document_id'])
        ET.SubElement(ae, "TYPE_DOCUMENT").text      = article['type_document']
        ET.SubElement(ae, "NUM_LIGNE").text          = str(article.get('num_ligne', ''))
        ET.SubElement(ae, "DESIGNATION").text        = article.get('designation', '')
        ET.SubElement(ae, "CODE_SH").text            = article.get('code_sh', '')
        ET.SubElement(ae, "QUANTITE").text           = str(q)
        ET.SubElement(ae, "UNITE").text              = article.get('unite', '')
        ET.SubElement(ae, "POIDS_NET_KG").text       = f"{article.get('poids_net', 0):.2f}"
        ET.SubElement(ae, "POIDS_BRUT").text         = f"{article.get('poids_brut', 0):.2f}"
        ET.SubElement(ae, "VALEUR_UNITAIRE").text    = f"{article.get('valeur_unitaire', 0):.2f}"
        ET.SubElement(ae, "VALEUR_TOTALE").text      = f"{v:.2f}"
        ET.SubElement(ae, "DEVISE").text             = financial_totals["devise"]
        ET.SubElement(ae, "ORIGINE").text            = article.get('origine', '')

    return minidom.parseString(ET.tostring(root, 'utf-8')).toprettyxml(indent="  ")


# ═══════════════════════════════════════════════════════════════════════════════
# Export Excel (XLSX) — tableaux encadrés, en-têtes bleus, lignes zébrées jaunes
# ═══════════════════════════════════════════════════════════════════════════════

# Palette
_FONT_NAME       = "Arial"
_BLUE_TITLE      = "1F4E78"   # bleu foncé — titres de section
_BLUE_HEADER     = "2E5395"   # bleu — en-têtes de tableau
_YELLOW_LABEL    = "FFE699"   # jaune — étiquettes clé/valeur
_YELLOW_ZEBRA    = "FFF2CC"   # jaune clair — lignes alternées
_YELLOW_TOTAL    = "FFD966"   # jaune vif — ligne de total
_WHITE           = "FFFFFF"


def _thin_border():
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _section_title(ws, row, title, span):
    """Ligne de titre de section : fond bleu, texte blanc gras, fusionnée et encadrée."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    border = _thin_border()
    for c in range(1, span + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", start_color=_BLUE_TITLE, end_color=_BLUE_TITLE)
        cell.border = border
    top_left = ws.cell(row=row, column=1)
    top_left.value = title
    top_left.font = Font(name=_FONT_NAME, bold=True, size=12, color=_WHITE)
    top_left.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _kv_row(ws, row, label, value):
    """Ligne clé/valeur : étiquette sur fond jaune, valeur sur fond blanc, encadrées."""
    border = _thin_border()
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = Font(name=_FONT_NAME, bold=True, size=10)
    lc.fill = PatternFill("solid", start_color=_YELLOW_LABEL, end_color=_YELLOW_LABEL)
    lc.border = border
    vc = ws.cell(row=row, column=2, value=value)
    vc.font = Font(name=_FONT_NAME, size=10)
    vc.fill = PatternFill("solid", start_color=_WHITE, end_color=_WHITE)
    vc.border = border
    vc.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _table_header(ws, row, headers):
    """En-tête de tableau : fond bleu, texte blanc gras, encadré."""
    border = _thin_border()
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name=_FONT_NAME, bold=True, size=10, color=_WHITE)
        cell.fill = PatternFill("solid", start_color=_BLUE_HEADER, end_color=_BLUE_HEADER)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return row + 1


def _table_row(ws, row, values, zebra_index):
    """Ligne de données : lignes paires en jaune clair, impaires en blanc, encadrées."""
    border = _thin_border()
    bg = _YELLOW_ZEBRA if zebra_index % 2 == 0 else _WHITE
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(name=_FONT_NAME, size=10)
        cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def prepare_xlsx_export(
    dossier,
    documents,
    articles,
    total_weight,
    total_value_doc,
    total_value_mad,
    financial_totals=None
):
    total_brut = sum(article.get('poids_brut', 0) for article in articles)
    first_doc  = documents[0] if documents else {}
    financial_totals = financial_totals or get_document_financial_totals(documents, articles)

    wb = Workbook()
    ws = wb.active
    ws.title = "Export Douanier"
    ws.sheet_view.showGridLines = False

    row = 1

    # ── DOSSIER ─────────────────────────────────────────────────────────────
    row = _section_title(ws, row, "INFORMATIONS DOSSIER", span=2)
    row = _kv_row(ws, row, "Référence",       dossier['ref_interne'])
    row = _kv_row(ws, row, "Statut",          dossier['statut'])
    row = _kv_row(ws, row, "Date création",   str(dossier['date_creation']))
    row = _kv_row(ws, row, "Numéro document", first_doc.get('numero_document', ''))
    row = _kv_row(ws, row, "Date document",   first_doc.get('date_document', ''))
    row = _kv_row(ws, row, "Devise",          financial_totals["devise"])
    row = _kv_row(ws, row, "Date export",     datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
    row = _kv_row(ws, row, "Générateur",      "DouanePro — Système de Transit Douanier")
    row += 1

    # ── EXPÉDITEUR ──────────────────────────────────────────────────────────
    row = _section_title(ws, row, "EXPÉDITEUR", span=2)
    row = _kv_row(ws, row, "Nom",         first_doc.get('expediteur_nom', ''))
    row = _kv_row(ws, row, "Adresse",     first_doc.get('expediteur_adresse', ''))
    row = _kv_row(ws, row, "Ville",       first_doc.get('expediteur_ville', ''))
    row = _kv_row(ws, row, "Code postal", first_doc.get('expediteur_code_postal', ''))
    row = _kv_row(ws, row, "Pays",        first_doc.get('expediteur_pays', ''))
    row += 1

    # ── DESTINATAIRE ────────────────────────────────────────────────────────
    row = _section_title(ws, row, "DESTINATAIRE", span=2)
    row = _kv_row(ws, row, "Nom",         first_doc.get('destinataire_nom', ''))
    row = _kv_row(ws, row, "Adresse",     first_doc.get('destinataire_adresse', ''))
    row = _kv_row(ws, row, "Ville",       first_doc.get('destinataire_ville', ''))
    row = _kv_row(ws, row, "Code postal", first_doc.get('destinataire_code_postal', ''))
    row = _kv_row(ws, row, "Pays",        first_doc.get('destinataire_pays', ''))
    row += 1

    # ── DOCUMENTS SOURCE ────────────────────────────────────────────────────
    doc_headers = ["ID Document", "Type", "Date import", "Montant HT", "TVA", "Total TTC", "Devise", "Chemin fichier"]
    row = _section_title(ws, row, "DOCUMENTS SOURCE", span=len(doc_headers))
    row = _table_header(ws, row, doc_headers)
    for i, doc in enumerate(documents):
        values = [
            doc['id'], doc['type_document'], str(doc['date_import']),
            round(_to_float(doc.get('montant_ht', 0)), 2),
            round(_to_float(doc.get('montant_tva', 0)), 2),
            round(_to_float(doc.get('valeur_totale_doc', 0)), 2),
            doc.get('devise', ''), doc['chemin_fichier'],
        ]
        row = _table_row(ws, row, values, i)
    row += 1

    # ── ARTICLES ────────────────────────────────────────────────────────────
    art_headers = [
        "N° Ligne", "Désignation", "Code SH", "Quantité", "Unité",
        "Poids Net (kg)", "Poids Brut (kg)",
        f"Valeur Unitaire ({financial_totals['devise']})",
        f"Valeur Totale ({financial_totals['devise']})",
        "Valeur Totale (MAD)", "Origine",
    ]
    row = _section_title(ws, row, "ARTICLES", span=len(art_headers))
    row = _table_header(ws, row, art_headers)
    for i, article in enumerate(articles):
        v_doc = article.get("valeur_devise", 0)
        v_mad = round(convert_currency(v_doc, financial_totals["devise"], "MAD"), 2)
        values = [
            article.get('num_ligne'),
            article.get('designation', ''),
            article.get('code_sh', '') or '—',
            article.get('quantite', 0),
            article.get('unite', ''),
            round(article.get('poids_net', 0), 3),
            round(article.get('poids_brut', 0), 3),
            round(article.get('valeur_unitaire', 0), 4),
            round(v_doc, 2),
            v_mad,
            article.get('origine', ''),
        ]
        row = _table_row(ws, row, values, i)
    row += 1

    # ── TOTAUX ──────────────────────────────────────────────────────────────
    row = _section_title(ws, row, "TOTAUX", span=2)
    row = _kv_row(ws, row, "Nombre de documents",   len(documents))
    row = _kv_row(ws, row, "Nombre d'articles",     len(articles))
    row = _kv_row(ws, row, "Poids net total (kg)",  round(total_weight, 3))
    row = _kv_row(ws, row, "Poids brut total (kg)", round(total_brut, 3))
    row = _kv_row(ws, row, "Montant HT",            round(financial_totals['montant_ht'], 2))
    row = _kv_row(ws, row, "TVA",                   round(financial_totals['montant_tva'], 2))

    # Ligne "Total TTC" mise en évidence en jaune vif
    border = _thin_border()
    lc = ws.cell(row=row, column=1, value="Total TTC")
    lc.font = Font(name=_FONT_NAME, bold=True, size=11)
    lc.fill = PatternFill("solid", start_color=_YELLOW_TOTAL, end_color=_YELLOW_TOTAL)
    lc.border = border
    vc = ws.cell(row=row, column=2, value=round(total_value_doc, 2))
    vc.font = Font(name=_FONT_NAME, bold=True, size=11)
    vc.fill = PatternFill("solid", start_color=_YELLOW_TOTAL, end_color=_YELLOW_TOTAL)
    vc.border = border
    row += 1

    row = _kv_row(ws, row, "Devise", financial_totals["devise"])

    # Ligne "Valeur totale (MAD)" mise en évidence en jaune vif
    lc = ws.cell(row=row, column=1, value="Valeur totale (MAD)")
    lc.font = Font(name=_FONT_NAME, bold=True, size=11)
    lc.fill = PatternFill("solid", start_color=_YELLOW_TOTAL, end_color=_YELLOW_TOTAL)
    lc.border = border
    vc = ws.cell(row=row, column=2, value=round(total_value_mad, 2))
    vc.font = Font(name=_FONT_NAME, bold=True, size=11)
    vc.fill = PatternFill("solid", start_color=_YELLOW_TOTAL, end_color=_YELLOW_TOTAL)
    vc.border = border
    row += 1

    # ── Largeurs de colonnes ────────────────────────────────────────────────
    widths = [22, 30, 16, 14, 14, 14, 16, 20, 22, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A1"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
def convert_currency(amount: float, from_currency: str, to_currency: str = 'USD') -> float:
    EXCHANGE_RATES = {'USD': 1.0, 'EUR': 1.08, 'MAD': 0.10}

    if from_currency == to_currency:
        return amount

    amount_in_usd = (
        amount * EXCHANGE_RATES.get(from_currency, 1.0)
        if from_currency != 'USD' else amount
    )

    return (
        amount_in_usd / EXCHANGE_RATES.get(to_currency, 1.0)
        if to_currency != 'USD' else amount_in_usd
    )